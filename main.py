import os
import json
import openpyxl as opxl

def execPgm():
    try:
        wb = opxl.load_workbook('sample.xlsx')
    except: 
        print("Didn't find the indicated file, create new one")
        # create a new one
        wb = opxl.Workbook()

    # print all the sheet names
    print("All sheets: {}".format(wb.sheetnames))

    # Get the active sheet
    # also could use wb['sheetname'] to get the active tab
    actSheet = wb.active
    print(actSheet.title)

    # change sheet name and color
    actSheet.title = "First sheet"
    actSheet.sheet_properties.tabColor = "1072BA"

    # Open json data and transfer data format from json to list[dict1, dict2, ...]
    jsonFilePath = './Database/salary.json'
    exists = os.path.isfile(jsonFilePath)
    if not exists:
        return
    f = open(jsonFilePath, encoding="utf-8")
    data = json.load(f)
    f.close()

    # get default sheet title and generate titles dict and append to excel
    print(data[0])
    print(data[0].keys())

    defaultSheetTitles = data[0].keys() #Notice!! The data[0] should have full keys, otherwise it will lead the lack of sheetTitles. 
    defaultSheetTitlesDict = {}

    for index, i in enumerate(defaultSheetTitles):
        tempKey = i
        tempVal = index + 1
        defaultSheetTitlesDict[tempKey] = tempVal
    print(defaultSheetTitlesDict)

    # insert titles at row1
    for index, x in enumerate(defaultSheetTitles):
        actSheet.cell(row = 1, column = index + 1, value = x)

    # insert the real data
    # loop every subData and insert every subData's element to dedicated site
    for i, subData in enumerate(data):
        for subTitles in defaultSheetTitles:
            # Avoid element lack in the json (Ref to ./Database/salary_lack.json)
            if (subTitles in subData):
                actSheet.cell(row = i + 2, column = defaultSheetTitlesDict[subTitles], value = subData[subTitles])

    

    wb.save('sample.xlsx')

if __name__ == "__main__":
    execPgm()